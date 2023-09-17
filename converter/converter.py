import numpy as np
import pandas as pd
from PIL import Image
import xlsxwriter
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
import time

time_now = time.localtime()
timestamp = time.strftime("%H%M%S", time_now)


class Converter:
    def __init__(self,
                 crossword_date         = None,
                 dark_mode              = None, 
                 #crossword_type         = None,
                 grid_size              = None
                 ) -> None:
        self.crossword_date = crossword_date
        self.dark_mode = dark_mode
        #self.crossword_type = crossword_type
        self.grid_size = grid_size
        self.crossword_image = self.open_crossword_image()
        self.crossword_width, self.crossword_height = self.crossword_image.size
        self.create_crossword_df()


    def open_crossword_image(self):
        filepath = f"PUT_SCREENSHOT_HERE\{self.crossword_date}.png"
        print(f'Reading the {self.crossword_date} crossword image...')
        return Image.open(filepath)
    

    def create_crossword_df(self) -> pd.DataFrame:
        self.crossword_df = pd.DataFrame(
            index=range(self.grid_size),
            columns=range(self.grid_size),
            dtype='bool'
            )
        
    
    def is_dark_mode_letter_box(self, rgb_value: tuple) -> bool:
        return True if all(val < 20 for val in rgb_value[:3]) else False


    def convert_crossword_screenshot_to_df(self) -> None:
        pix = self.crossword_image.load()

        cell_size = (self.crossword_height + self.crossword_width)/(2*self.grid_size)
        x_pixels = np.linspace(cell_size/2, self.crossword_height - cell_size/2, self.grid_size)
        y_pixels = np.linspace(cell_size/2, self.crossword_width - cell_size/2, self.grid_size)

        for x_idx, x_pixel in enumerate(x_pixels):
            for y_idx, y_pixel in enumerate(y_pixels):
                
                x_pixel = np.floor(x_pixel)
                y_pixel = np.floor(y_pixel)
                rgb_value = pix[x_pixel, y_pixel]

                if not self.dark_mode:
                    is_letter_box = (rgb_value[:3] == (255, 255, 255))
                else:
                    is_letter_box = self.is_dark_mode_letter_box(rgb_value)

                if is_letter_box:
                    self.crossword_df.at[x_idx, y_idx] = True
                    cell_colour = 'White'
                else:
                    self.crossword_df.at[x_idx, y_idx] = False
                    cell_colour = 'Black'

                # Show cell colour information
                print(f'Color at cell ({x_idx}, {y_idx}), co-ordinate ({x_pixel}, {y_pixel}): {rgb_value} -> {cell_colour}')

        # Put df in correct order
        # self.crossword_df = self.crossword_df.loc[::-1]
        self.crossword_df = self.crossword_df.T

        return
    

    def create_crossword_xlsx(self):

        output_filename = f'outputs\{self.crossword_date} {timestamp}.xlsx'
        writer = pd.ExcelWriter(output_filename, engine='xlsxwriter')
        self.crossword_df.to_excel(writer, sheet_name='Sheet1', index=False)
        writer.close()

        # Format xlsx
        wb = openpyxl.load_workbook(output_filename)
        sheet = wb['Sheet1']
        sheet.delete_rows(1, amount=1)

        # Cell fills and borders
        black_fill = PatternFill(start_color='00000000', end_color='00000000', fill_type='solid')
        thin_border = Side(border_style="thin", color="000000")
        
        # Iterate through all cells in the xlsx
        for row in sheet.iter_rows(min_row=1, min_col=1, max_row=self.grid_size, max_col=self.grid_size):
            for cell in row:
                
                # Resize cell
                sheet.row_dimensions[cell.row].height = 31.5
                sheet.column_dimensions[cell.column_letter].width = 40/7

                # Colour cell
                if str(cell.value).lower() == "false":
                    cell.fill = black_fill

                # Add borders
                cell.border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)

                # Clear cell text
                cell.value = None
        
        # Save the modified workbook
        wb.save(output_filename)
        print(f"Crossword converted. Saved in putputs folder as as {self.crossword_date} {timestamp}.xlsx.")
